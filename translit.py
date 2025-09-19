# Сохраните этот код как translit_openpyxl.py

import openpyxl
from unidecode import unidecode
import glob
import os

def transliterate_text(text):
    """Транслитерирует текст с использованием unidecode."""
    if text is None:
        return None
    return unidecode(str(text))

def process_excel_file(file_path, column_to_process):
    """Обрабатывает один Excel-файл, транслитерируя указанный столбец."""
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except Exception as e:
        print(f"Ошибка при чтении файла {file_path}: {e}")
        return

    # Найти индекс столбца по названию
    headers = [cell.value for cell in sheet[1]]
    try:
        col_index = headers.index(column_to_process) + 1
    except ValueError:
        print(f" - Столбец '{column_to_process}' не найден в файле {os.path.basename(file_path)}. Файл пропущен.")
        return

    # Создать новый столбец для транслитерированного текста
    new_col_index = sheet.max_column + 1
    sheet.cell(row=1, column=new_col_index, value=f"{column_to_process}_translit")

    # Пройти по строкам и транслитерировать текст
    for row_num in range(2, sheet.max_row + 1):
        original_text = sheet.cell(row=row_num, column=col_index).value
        transliterated_text = transliterate_text(original_text)
        sheet.cell(row=row_num, column=new_col_index, value=transliterated_text)

    print(f" - Столбец '{column_to_process}' успешно транслитерирован в новый столбец '{sheet.cell(row=1, column=new_col_index).value}'.")
    
    # Сохранение результата в новый файл
    output_path = f"translit_{os.path.basename(file_path)}"
    try:
        workbook.save(output_path)
        print(f"Результат сохранен в файл: {output_path}\n")
    except Exception as e:
        print(f"Ошибка при сохранении файла {output_path}: {e}\n")


if __name__ == "__main__":
    print("Поиск Excel-файлов (.xlsx) в текущей папке...")
    
    excel_files = glob.glob("*.xlsx")
    
    if not excel_files:
        print("В текущей папке не найдено ни одного файла .xlsx.")
        input("Нажмите Enter, чтобы выйти.")
    else:
        # Показать пользователю столбцы из первого найденного файла
        first_file_path = excel_files[0]
        try:
            workbook = openpyxl.load_workbook(first_file_path)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            print(f"\nСтолбцы в файле '{os.path.basename(first_file_path)}':")
            for col_name in headers:
                print(f" - {col_name}")
        except Exception as e:
            print(f"Не удалось прочитать столбцы из файла '{os.path.basename(first_file_path)}': {e}")
            input("Нажмите Enter, чтобы выйти.")
            exit()
            
        # Запросить у пользователя название столбца
        column_name = input("\nВведите название столбца, который нужно обработать (точно как в списке): ").strip()
        
        if not column_name:
            print("Название столбца не было введено. Процесс отменен.")
        else:
            print("\nНачинаем обработку...")
            for file in excel_files:
                print(f"\nОбработка файла: {file}")
                process_excel_file(file, column_name)
        
        print("\nПроцесс завершен.")
        input("Нажмите Enter, чтобы выйти.")
        